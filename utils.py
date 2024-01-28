import matplotlib.pyplot as plt
import numpy
from openpyxl import load_workbook


def read_worksheet(filename, sheet):
    assert isinstance(sheet, str)
    return load_workbook(filename=filename)[sheet]


def read_all_worksheets(filename, except_sheet='', smiles_column=''):
    wb = load_workbook(filename=filename)
    descriptors_name = []
    descriptors_list = []
    smiles_list = []
    float_conversion_failed_rows = set()

    for idx, sheet in enumerate(wb.worksheets):
        print(idx, "/", len(wb.worksheets), " sheet")
        if sheet.title == except_sheet:
            continue
        for col in sheet.iter_cols():
            descriptor_column = []
            is_smiles_column = False

            for cell in col:
                if cell.row == 1:
                    if cell.value == smiles_column:
                        is_smiles_column = True
                    else:
                        descriptors_name.append(cell.value)
                else:
                    try:
                        if is_smiles_column:
                            smiles_list.append(cell.value)
                        else:
                            cell_value = float(cell.value)
                            descriptor_column.append(cell_value)
                    except:
                        # Excel indexation starts at 1 and pythons at 0.
                        # Headers are occupying 1st row, therefore "cell.row - 2"
                        float_conversion_failed_rows.add(cell.row - 2)
                        # Add dummy data so we keep the correct structure, later we delete the whole row
                        descriptor_column.append(0)

            if len(descriptor_column) != 0:
                descriptors_list.append(descriptor_column)

    # Check 2D array row-col consistency
    num_of_rows = set()
    if smiles_list:
        num_of_rows.add(len(smiles_list))
    for col in descriptors_list:
        num_of_rows.add(len(col))
    if len(num_of_rows) > 1:
        raise "Some columns have different number of rows " + str(num_of_rows)

    # Delete rows with invalid float values. Delete rows in SMILES list as well
    for row_index in sorted(float_conversion_failed_rows, reverse=True):
        for col in descriptors_list:
            del col[row_index]
        if smiles_list:
            del smiles_list[row_index]

    if smiles_list:
        return descriptors_name, descriptors_list, smiles_list
    else:
        return descriptors_name, descriptors_list


def get_one_row_list_from_range(worksheet_range):
    my_list = []
    for row in worksheet_range:
        row_list = [cell.value for cell in list(row)]
        my_list.append(row_list[0])
    return my_list


def calc_r_squared(actual, predict):
    corr_matrix = numpy.corrcoef(actual, predict)
    corr = corr_matrix[0, 1]
    r_sq = corr ** 2

    return r_sq


def do_plot(actual, predict, test_actual, test_predicted, color, path=''):
    # Draw a trendline
    z = numpy.polyfit(actual, predict, 1)
    p = numpy.poly1d(z)
    plt.plot(actual, p(actual), "k-", lw=1)
    plt.scatter(actual, predict, color=color, marker='s', edgecolors="black")
    plt.scatter(test_actual, test_predicted, color=color, marker='x')
    plt.xticks(fontsize=14)
    plt.yticks(fontsize=14)
    plt.xlabel("RI (actual value)", fontsize=18)
    plt.ylabel("RI (predicted value)", fontsize=18)

    if path != '':
        plt.savefig(path, bbox_inches='tight', dpi=350)
        plt.clf()


def r_squared(a, b):
    corr_matrix = numpy.corrcoef(a, b)
    corr = corr_matrix[0, 1]
    return corr ** 2


def regressionFeaturesFilter(y, X, X_names, threshold):
    X_t = numpy.array(X).T
    filtered_X_t = []
    filtered_names = []
    for i, X in enumerate(X_t):
        if r_squared(y, X_t[i]) > threshold:
            filtered_names.append(X_names[i])
            filtered_X_t.append(X_t[i])
    return numpy.array(filtered_X_t).T, filtered_names
